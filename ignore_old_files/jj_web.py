#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
jj_web.py
抓取晋江作品库某 1 年作品（默认前 N 页），字段：
作者、作品（带超链接）、类型、进度、字数、作品积分、发表时间
用于生成jjwxc_2025.xlsx

特性：
- 断点续爬：每条数据落盘 JSONL（带 _page 标记），中断后可继续
- 防崩溃：单页失败不终止；保存失败 HTML 到 debug_html/；记录失败页到 failed_pages.txt
- 更“像浏览器”的请求头 + 友好延时 + 重试
- 最终导出 Excel（作品列超链接）

使用示例：
1) 先试爬 5 页：
   python jj_web.py --end-page 5

2) 正式爬 50 页：
   python jj_web.py --end-page 50

3) 仅重试失败页：
   python jj_web.py --retry-failed

4) 导出 Excel（不爬，只把现有 jsonl 导出）：
   python jj_web.py --export-only

可选：如果你遇到反爬验证/需要登录（晋江一般读10页限制一次），可把浏览器 Cookie 粘进 --cookie
   python jj_web.py --end-page 5 --cookie "你的整段Cookie字符串"

   比如：python jj_web.py --end-page 50 --cookie "$(cat cookie.txt)" --min-delay 1 --max-delay 3
   目前同目录的cookie.txt就是我从浏览器登陆自己的晋江账号复制出来的。
"""

import argparse
import json
import os
import random
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

#可替换成不同年份的网页
BASE_URL = "https://www.jjwxc.net/bookbase.php"

# 你给的第一页参数 + 第二页出现的默认参数（更稳）
BASE_PARAMS = {
    "fw0": "0",
    "fbsj2022": "2022",
    "novelbefavoritedcount0": "0",
    "yc0": "0",
    "xx0": "0",
    "mainview0": "0",
    "sd0": "0",
    "lx0": "0",
    "collectiontypes": "ors",
    "notlikecollectiontypes": "ors",
    "bq": "-1",
    "removebq": "",
    "searchkeywords": "",
    "sortType": "0",
    "isfinish": "0",
}

# 更真实的请求头
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Referer": BASE_URL,
}


# ---------------------------
# 网络层
# ---------------------------
def robust_get(
    session: requests.Session,
    params: dict,
    retries: int = 6,
    timeout: int = 25,
) -> str:
    last_err = None
    for i in range(retries):
        try:
            r = session.get(BASE_URL, params=params, headers=HEADERS, timeout=timeout)
            r.raise_for_status()

            # 晋江常见编码：GBK/GB2312；requests 有时会猜错
            if not r.encoding or r.encoding.lower() == "iso-8859-1":
                r.encoding = r.apparent_encoding or "gbk"
            return r.text
        except Exception as e:
            last_err = e
            # 递增退避
            time.sleep(2 + i * 2 + random.random())
    raise RuntimeError(f"请求失败，重试 {retries} 次仍失败：{last_err}")


def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


# ---------------------------
# 解析层
# ---------------------------
def extract_total_pages_m_p(html: str) -> Optional[str]:
    """
    从页面中提取 m_p（总页数），用于稳定翻页。
    """
    soup = BeautifulSoup(html, "lxml")

    inp = soup.find("input", attrs={"name": "m_p"})
    if inp and inp.get("value"):
        return inp["value"]

    # 兼容“共 N 页”形式
    text = soup.get_text(" ", strip=True)
    m = re.search(r"共\s*(\d+)\s*页", text)
    if m:
        return m.group(1)

    return None


def looks_like_blocked(html: str) -> bool:
    """
    粗略判断是否被风控/验证码/非正常页面（用于更友好提示，不是硬判定）。
    """
    t = html.lower()
    keywords = [
        "验证码",
        "安全验证",
        "访问受限",
        "forbidden",
        "403",
        "验证",
        "robot",
        "anti",
        "waf",
        "登录",
        "login",
    ]
    return any(k in t for k in keywords)


def find_target_table(soup: BeautifulSoup):
    """
    尽量稳健地定位包含目标列的表格。
    """
    candidates = soup.find_all("table")
    for table in candidates:
        header_text = clean_text(table.get_text(" "))
        if all(k in header_text for k in ["作者", "作品", "类型", "进度", "字数", "作品积分", "发表时间"]):
            return table
    return None


def parse_rows(html: str) -> List[Dict]:
    soup = BeautifulSoup(html, "lxml")
    table = find_target_table(soup)

    if not table:
        raise ValueError("未找到目标表格（可能触发反爬/需要登录/或页面结构变化）")

    rows: List[Dict] = []
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue

        author = clean_text(tds[0].get_text(" "))

        a = tds[1].find("a")
        work_title = clean_text(a.get_text(" ")) if a else clean_text(tds[1].get_text(" "))
        work_href = urljoin(BASE_URL, a["href"]) if a and a.get("href") else ""

        category = clean_text(tds[2].get_text(" "))
        progress = clean_text(tds[3].get_text(" "))
        word_count = clean_text(tds[4].get_text(" "))
        score = clean_text(tds[5].get_text(" "))
        pub_time = clean_text(tds[6].get_text(" "))

        # 跳过表头行/空行
        if author == "作者" and work_title == "作品":
            continue
        if not author or not work_title:
            continue

        rows.append(
            {
                "作者": author,
                "作品": work_title,
                "作品链接": work_href,
                "类型": category,
                "进度": progress,
                "字数": word_count,
                "作品积分": score,
                "发表时间": pub_time,
            }
        )
    return rows


def build_params(page: int, m_p: Optional[str]) -> dict:
    params = dict(BASE_PARAMS)
    params["page"] = str(page)
    if m_p:
        params["m_p"] = str(m_p)
    return params


def fetch_page_html(session: requests.Session, page: int, m_p: Optional[str]) -> str:
    params = build_params(page, m_p)
    return robust_get(session, params)


# ---------------------------
# 断点续爬（JSONL）
# ---------------------------
def load_done_pages(jsonl_path: Path) -> Set[int]:
    done: Set[int] = set()
    if not jsonl_path.exists():
        return done
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except Exception:
                continue
            p = obj.get("_page")
            if isinstance(p, int):
                done.add(p)
    return done


def append_page_rows(jsonl_path: Path, page: int, rows: List[Dict]):
    with jsonl_path.open("a", encoding="utf-8") as f:
        for r in rows:
            r2 = dict(r)
            r2["_page"] = page
            f.write(json.dumps(r2, ensure_ascii=False) + "\n")


def load_all_rows(jsonl_path: Path) -> List[Dict]:
    rows: List[Dict] = []
    if not jsonl_path.exists():
        return rows
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            obj.pop("_page", None)
            rows.append(obj)
    return rows


# ---------------------------
# 失败页记录
# ---------------------------
def load_failed_pages(failed_log: Path) -> List[int]:
    if not failed_log.exists():
        return []
    pages = []
    with failed_log.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            # 格式：page\tError: msg
            try:
                p = int(line.split("\t", 1)[0])
                pages.append(p)
            except Exception:
                continue
    # 去重保持顺序
    seen = set()
    uniq = []
    for p in pages:
        if p not in seen:
            uniq.append(p)
            seen.add(p)
    return uniq


def log_failed(failed_log: Path, page: int, err: Exception):
    with failed_log.open("a", encoding="utf-8") as f:
        f.write(f"{page}\t{type(err).__name__}: {err}\n")


# ---------------------------
# 导出 Excel
# ---------------------------
def save_to_xlsx(rows: List[Dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "JJWXC_2022"

    headers = ["作者", "作品", "类型", "进度", "字数", "作品积分", "发表时间"]
    ws.append(headers)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(
            [
                r.get("作者", ""),
                r.get("作品", ""),
                r.get("类型", ""),
                r.get("进度", ""),
                r.get("字数", ""),
                r.get("作品积分", ""),
                r.get("发表时间", ""),
            ]
        )
        row_idx = ws.max_row
        link = r.get("作品链接", "")
        if link:
            c = ws.cell(row=row_idx, column=2)
            c.hyperlink = link
            c.font = Font(color="0000EE", underline="single")

    # 列宽（简单估算）
    for col in range(1, len(headers) + 1):
        max_len = 10
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(row=rr, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    wb.save(out_path)


# ---------------------------
# 主流程
# ---------------------------
def crawl_pages(
    end_page: int,
    jsonl_path: Path,
    debug_dir: Path,
    failed_log: Path,
    cookie: Optional[str],
    min_delay: float,
    max_delay: float,
    fail_sleep_min: float,
    fail_sleep_max: float,
    export_xlsx: Path,
    export_each_run: bool,
):
    debug_dir.mkdir(exist_ok=True)

    done_pages = load_done_pages(jsonl_path)
    print(f"📌 已完成页数：{len(done_pages)}（断点续爬将跳过这些页）")

    with requests.Session() as session:
        if cookie:
            session.headers.update({"Cookie": cookie})

        # 先抓第一页，用于解析 m_p（总页数）
        first_params = dict(BASE_PARAMS)
        first_params["page"] = "1"
        first_html = robust_get(session, first_params)

        m_p = extract_total_pages_m_p(first_html)

        # 需要爬的页
        pages_to_crawl = [p for p in range(1, end_page + 1) if p not in done_pages]
        if not pages_to_crawl:
            print("✅ 目标页范围内都已爬完。")
        else:
            for p in tqdm(pages_to_crawl, desc="Crawling pages (resume)"):
                html = ""
                try:
                    if p == 1:
                        html = first_html
                    else:
                        html = fetch_page_html(session, p, m_p=m_p)

                    # 有些失败是“正常返回但不是表格页”
                    if looks_like_blocked(html):
                        # 不直接判死刑，但提示会更清晰
                        # 仍然交给 parse_rows 决定是否能解析
                        pass

                    rows = parse_rows(html)

                    # 正常每页应接近 100
                    if len(rows) < 50:
                        raise RuntimeError(f"解析条数异常偏少：{len(rows)}（可能风控/结构变化/参数异常）")

                    append_page_rows(jsonl_path, p, rows)

                    # 礼貌延时（建议 2~5 秒起）
                    time.sleep(random.uniform(min_delay, max_delay))

                except Exception as e:
                    # 保存失败 html
                    html_path = debug_dir / f"page_{p}_failed.html"
                    try:
                        html_path.write_text(html or "", encoding="utf-8", errors="ignore")
                    except Exception:
                        pass

                    log_failed(failed_log, p, e)

                    print(
                        f"\n⚠️ 第 {p} 页失败：{type(e).__name__}: {e}\n"
                        f"   - 已保存：{html_path}\n"
                        f"   - 已记录：{failed_log}\n"
                        f"   - 将继续下一页...\n"
                    )

                    # 失败后多休息一会儿
                    time.sleep(random.uniform(fail_sleep_min, fail_sleep_max))
                    continue

    if export_each_run:
        rows = load_all_rows(jsonl_path)
        print(f"📦 当前累计条目：{len(rows)}，开始导出 Excel...")
        save_to_xlsx(rows, export_xlsx)
        print(f"✅ 导出完成：{export_xlsx}")


def export_only(jsonl_path: Path, export_xlsx: Path):
    rows = load_all_rows(jsonl_path)
    print(f"📦 读取到 {len(rows)} 条记录，导出 Excel...")
    save_to_xlsx(rows, export_xlsx)
    print(f"✅ 导出完成：{export_xlsx}")


def retry_failed(
    jsonl_path: Path,
    debug_dir: Path,
    failed_log: Path,
    cookie: Optional[str],
    min_delay: float,
    max_delay: float,
    fail_sleep_min: float,
    fail_sleep_max: float,
    export_xlsx: Path,
):
    debug_dir.mkdir(exist_ok=True)

    failed_pages = load_failed_pages(failed_log)
    if not failed_pages:
        print("✅ failed_pages.txt 为空，无需重试。")
        return

    # 只重试那些尚未成功写入的页（避免重复）
    done_pages = load_done_pages(jsonl_path)
    targets = [p for p in failed_pages if p not in done_pages]
    if not targets:
        print("✅ 失败页都已经在 jsonl 里成功写入过了，无需重试。")
        return

    print(f"🔁 将重试失败页（未完成的）：{targets}")

    with requests.Session() as session:
        if cookie:
            session.headers.update({"Cookie": cookie})

        # 为了拿 m_p
        first_params = dict(BASE_PARAMS)
        first_params["page"] = "1"
        first_html = robust_get(session, first_params)
        m_p = extract_total_pages_m_p(first_html)

        for p in tqdm(targets, desc="Retry failed pages"):
            html = ""
            try:
                html = first_html if p == 1 else fetch_page_html(session, p, m_p=m_p)
                rows = parse_rows(html)
                if len(rows) < 50:
                    raise RuntimeError(f"解析条数异常偏少：{len(rows)}")

                append_page_rows(jsonl_path, p, rows)
                time.sleep(random.uniform(min_delay, max_delay))

            except Exception as e:
                html_path = debug_dir / f"page_{p}_retry_failed.html"
                try:
                    html_path.write_text(html or "", encoding="utf-8", errors="ignore")
                except Exception:
                    pass

                print(
                    f"\n⚠️ 重试仍失败：第 {p} 页：{type(e).__name__}: {e}\n"
                    f"   - 已保存：{html_path}\n"
                )
                time.sleep(random.uniform(fail_sleep_min, fail_sleep_max))
                continue

    # 重试后导出
    rows = load_all_rows(jsonl_path)
    print(f"📦 当前累计条目：{len(rows)}，开始导出 Excel...")
    save_to_xlsx(rows, export_xlsx)
    print(f"✅ 导出完成：{export_xlsx}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--end-page", type=int, default=5, help="爬取到第几页（包含）。先试5页建议。")
    parser.add_argument("--jsonl", type=str, default="jjwxc_2025_pages.jsonl", help="断点续爬缓存 JSONL 文件名")
    parser.add_argument("--export", type=str, default="jjwxc_2025.xlsx", help="导出 Excel 文件名")
    parser.add_argument("--cookie", type=str, default=None, help="可选：浏览器复制的 Cookie 字符串（提高成功率）")

    parser.add_argument("--min-delay", type=float, default=1.0, help="每页成功后最小延时（秒）")
    parser.add_argument("--max-delay", type=float, default=3.0, help="每页成功后最大延时（秒）")
    parser.add_argument("--fail-sleep-min", type=float, default=6.0, help="单页失败后最小休眠（秒）")
    parser.add_argument("--fail-sleep-max", type=float, default=12.0, help="单页失败后最大休眠（秒）")

    parser.add_argument("--export-only", action="store_true", help="不爬取，只导出现有 JSONL 到 Excel")
    parser.add_argument("--retry-failed", action="store_true", help="仅重试 failed_pages.txt 里失败的页（未完成的）")
    parser.add_argument("--no-export", action="store_true", help="跑完不导出 Excel（只更新 JSONL）")

    args = parser.parse_args()

    jsonl_path = Path(args.jsonl)
    export_xlsx = Path(args.export)
    debug_dir = Path("debug_html")
    failed_log = Path("failed_pages.txt")

    if args.export_only:
        export_only(jsonl_path, export_xlsx)
        return

    if args.retry_failed:
        retry_failed(
            jsonl_path=jsonl_path,
            debug_dir=debug_dir,
            failed_log=failed_log,
            cookie=args.cookie,
            min_delay=args.min_delay,
            max_delay=args.max_delay,
            fail_sleep_min=args.fail_sleep_min,
            fail_sleep_max=args.fail_sleep_max,
            export_xlsx=export_xlsx,
        )
        return

    crawl_pages(
        end_page=args.end_page,
        jsonl_path=jsonl_path,
        debug_dir=debug_dir,
        failed_log=failed_log,
        cookie=args.cookie,
        min_delay=args.min_delay,
        max_delay=args.max_delay,
        fail_sleep_min=args.fail_sleep_min,
        fail_sleep_max=args.fail_sleep_max,
        export_xlsx=export_xlsx,
        export_each_run=(not args.no_export),
    )


if __name__ == "__main__":
    main()
