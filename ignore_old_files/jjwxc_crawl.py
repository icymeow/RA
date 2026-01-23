import re
import time
import random
from typing import Optional, Dict, List
from pathlib import Path

import requests
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm

#========= 此code用来从 jjwxc_2025.xlsx 里爬取“文案”“内容标签”“一句话简介”“立意”等字段 ========#

# ========= 配置 =========
INPUT_XLSX = "jjranking_10years.xlsx"  # 如果你在本机跑，放同目录；如果在你的环境里是 /mnt/data/jjranking.xlsx 就改成那个

OUTPUT_XLSX = "jjwxc_10yrs_withtags.xlsx"
OUTPUT_CSV = "jjwxc_10yrs_withtags.csv"


SLEEP_MIN = 0.3
SLEEP_MAX = 0.9
RETRIES = 2
TIMEOUT = 12

DEBUG_SNIPPET = False

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://www.jjwxc.net/",
}

# Excel 不允许的控制字符（0x00-0x08,0x0B-0x0C,0x0E-0x1F）
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

# ========= ✅ 内容标签清洗 =========

VALID_TAGS = {
    "甜文","情有独钟","爽文","穿越时空","天作之合","强强","穿书","天之骄子","系统","成长","豪门世家","都市","日常","宫廷侯爵",
    "种田文","重生","仙侠修真","年代文","业界精英","破镜重圆","先婚后爱","升级流","娱乐圈","万人迷","快穿","灵异神怪","无限流",
    "幻想空间","校园","励志","基建","救赎","群像","沙雕","ABO","美食","欢喜冤家","年下","治愈","现代架空","末世","综漫",
    "追爱火葬场","星际","HE","团宠","青梅竹马","逆袭","异能","直播","暗恋","因缘邂逅","生子","萌宠","悬疑推理","美强惨",
    "高岭之花","囤货","相爱相杀","正剧","市井生活","少年漫","朝堂","西方罗曼","轻松","咒回","经营","打脸","历史衍生","柯南",
    "英美衍生","女强","荒野求生","惊悚","狗血","未来架空","女配","体育竞技","克苏鲁","抽奖抽卡","文野","三教九流","钓系","玄学",
    "近水楼台","迪化流","婚恋","单元文","马甲文","游戏网游","超级英雄","开挂","脑洞","花季雨季","布衣生活","科幻","萌娃","东方玄幻",
    "西幻","清穿","白月光","废土","爆笑","复仇虐渣","异世大陆","边缘恋歌","反套路","恋爱合约","古代幻想","日久生情","虫族","江湖",
    "论坛体","机甲","家教","鬼灭","女扮男装","魔幻","火影","科举","排球少年","乙女向","第四天灾","天选之子","阴差阳错","随身空间",
    "足球","网王","电竞","平步青云","日韩泰","龙傲天","忠犬","港风","前世今生","综艺","宅斗","赛博朋克","武侠","创业","热血","田园",
    "制服情缘","全息","规则怪谈","古穿今","失忆","腹黑","海贼王","真假少爷","御姐","权谋","宫斗","虐文","炮灰","宋穿","学霸","灵魂转换",
    "异想天开","读心术","都市异闻","咸鱼","师徒","乔装改扮","吐槽","异闻传说","时代奇缘","古典名著","剧透","姐弟恋","唐穿","史诗奇幻",
    "汉穿","哨向","男配","红楼梦","猎人","对照组","职场","时代新风","卡牌","赶山赶海","刀剑乱舞","多重人格","真假千金","现实","明穿","燃",
    "弹幕","西方名著","签到流","吐槽役","星穹铁道","秦穿","灵气复苏","总裁","位面","神话传说","转生","预知","女尊","原神","黑篮","神豪流",
    "三国穿越","高智商","犬夜叉","公路文","非遗","NPC","纸片人","少女漫","民国","大冒险","时尚圈","剑网3","群穿","毒舌","冰山","国风幻想",
    "模拟器","读档流","性别转换","FGO","傲娇","替身","烧脑","召唤流","商战","美娱","极品亲戚","吃货","封神","洪荒","开荒","奇谭","七五",
    "app","漫穿","JOJO","银魂","齐神","蓝锁","网红","暖男","萌","中二","聊斋","骑士与剑","血族","中世纪","亡灵异族","原始社会","恶役","御兽",
    "七年之痒","天降","盲盒","魔法少女","蒸汽朋克","锦鲤","扶贫","亚人","特摄","交换人生","魔王勇者","BE","死神","悲剧","红包群","网配","曲艺",
    "对话体","港台","SD","婆媳","圣斗士","绝区零"
}


# 常见分隔符：空格/全角空格/标点/斜杠/竖线/#/换行等
_TAG_SEP_RE = re.compile(r"[、，,;；/|｜#\u3000\s]+")
# 去掉括号/书名号/特殊包裹符号
_TAG_STRIP_RE = re.compile(r"^[【\[\(（<《『「]+|[】\]\)）>》』」]+$")

def clean_tags(raw: Optional[str], valid_tags: set) -> Dict[str, Optional[object]]:
    """
    输入：原始“内容标签”字符串（可能很脏）
    输出：
      - 内容标签_clean: 用空格拼好的干净标签串
      - 内容标签_list: 干净标签 list（保持出现顺序、去重）
      - 内容标签_raw: 原始字符串（方便你排查）
      - 内容标签_unknown: 不在 valid_tags 里的“疑似标签”list（可选，用于扩充词表）
    """
    if not raw:
        return {
            "内容标签_raw": raw,
            "内容标签_clean": None,
            "内容标签_list": None,
            "内容标签_unknown": None,
        }

    s = raw.strip()

    # 一些常见噪声词（如果你抓到过可以继续加）
    noise_phrases = [
        "标签", "内容标签", "更多搜索", "点击", "收藏", "评论", "作者", "作品简评",
        "文章基本信息", "主角", "配角", "立意", "一句话简介"
    ]
    for w in noise_phrases:
        s = s.replace(w, " ")

    # 统一符号（全角冒号等）
    s = s.replace("：", " ").replace(":", " ")
    # 有些页面会出现 “#甜文#” 这种，先把#当分隔
    s = s.replace("#", " ")

    # 先粗切分
    parts = [p for p in _TAG_SEP_RE.split(s) if p]

    cleaned = []
    unknown = []
    seen = set()

    for p in parts:
        p = p.strip()
        # 去掉首尾包裹符号
        p = _TAG_STRIP_RE.sub("", p).strip()
        if not p:
            continue

        # 关键：只保留你词表里的合法标签
        if p in valid_tags:
            if p not in seen:
                cleaned.append(p)
                seen.add(p)
        else:
            # 不是合法标签：先记录，后续你想扩充词表很有用
            # 也可以选择直接不记录，把 unknown 相关逻辑删掉
            if p not in seen:
                unknown.append(p)
                seen.add(p)

    return {
        "内容标签_raw": raw,
        "内容标签_clean": " ".join(cleaned) if cleaned else None,
        "内容标签_list": cleaned if cleaned else None,
        "内容标签_unknown": unknown if unknown else None,
    }


def sleep_a_bit():
    time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))


def fetch_html(session: requests.Session, url: str) -> str:
    last_err = None
    for _ in range(RETRIES):
        try:
            resp = session.get(url, headers=HEADERS, timeout=TIMEOUT)

            # 编码容错：晋江经常是 GBK/GB18030
            if not resp.encoding or resp.encoding.lower() in ("iso-8859-1", "ascii"):
                resp.encoding = resp.apparent_encoding or "gb18030"

            if resp.status_code == 200 and resp.text:
                return resp.text

            last_err = f"HTTP {resp.status_code}"
        except Exception as e:
            last_err = repr(e)

        sleep_a_bit()

    raise RuntimeError(f"Failed to fetch {url}: {last_err}")


def normalize(s: str) -> str:
    s = re.sub(r"\r", "", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def strip_illegal_xlsx_chars(x):
    if isinstance(x, str):
        return _ILLEGAL_XLSX_RE.sub("", x)
    return x

# ========= ✅ 新方法：纯文本切片抽字段 =========
def extract_fields_from_text(full_text: str) -> Dict[str, Optional[str]]:
    t = normalize(full_text)

    out = {
        "文案": None,
        "内容标签": None,
        "内容标签_list": None,
        "一句话简介": None,
        "立意": None,
    }

    lines = [x.strip() for x in t.split("\n")]

    # 下一个字段的“标题”出现时就停止收集
    STOP_LABELS = {
        "文案", "内容标签", "一句话简介", "立意",
        "主角", "配角", "其他", "其它", "角色",
        "作品简评", "文章基本信息", "作者", "评论",
        "所属系列", "文章进度", "全文字数", "是否出版",
        "主角视角", "配角视角"
    }

    def is_label_line(s: str) -> bool:
        s2 = s.strip()
        if not s2:
            return False
        # 形如 “内容标签：” 或 “立意：” 或 “内容标签”
        head = s2.split("：", 1)[0].split(":", 1)[0].strip()
        return head in STOP_LABELS

    def pick_block(label: str) -> Optional[str]:
        """
        从 label 行开始，收集 label 后面的内容（可能在同一行，也可能在后续多行）
        直到遇到另一个字段标题行停止。
        """
        for i, line in enumerate(lines):
            # 支持 “label：xxx” / “label:xxx” / “label”
            if line == label or line.startswith(label + "：") or line.startswith(label + ":"):
                vals = []

                # 同一行里冒号后的内容
                if "：" in line:
                    tail = line.split("：", 1)[1].strip()
                    if tail:
                        vals.append(tail)
                elif ":" in line:
                    tail = line.split(":", 1)[1].strip()
                    if tail:
                        vals.append(tail)

                # 下一行开始继续收集（很多标签就是一行一个）
                j = i + 1
                while j < len(lines):
                    nxt = lines[j].strip()
                    if not nxt:
                        j += 1
                        continue
                    if is_label_line(nxt) and not (nxt == label):  # 遇到下一个字段
                        break
                    vals.append(nxt)
                    j += 1

                # 拼接成一串（用空格分隔）
                joined = " ".join([v for v in vals if v]).strip()
                return joined if joined else None
        return None

    # ✅ 多行字段抓取
    out["内容标签"] = pick_block("内容标签")
    out["一句话简介"] = pick_block("一句话简介")
    out["立意"] = pick_block("立意")

    # ✅ 内容标签清洗：只保留 VALID_TAGS
    tag_pack = clean_tags(out["内容标签"], VALID_TAGS)

    # 如果你想保留原字段名“内容标签”=干净结果：
    out["内容标签"] = tag_pack["内容标签_clean"]
    out["内容标签_list"] = tag_pack["内容标签_list"]

    # 可选：多加两列方便排查/扩充词表
    out["内容标签_raw"] = tag_pack["内容标签_raw"]
    out["内容标签_unknown"] = tag_pack["内容标签_unknown"]


    # ✅ 文案：从文案开始，到内容标签出现前结束（你的需求）
    start_pat = r"(?:^|\n)\s*[【\[]?\s*文案\s*[】\]]?\s*[:：]?\s*\n?"
    m = re.search(start_pat, t)
    if not m:
        return out

    start = m.end()
    m_stop = re.search(r"(?:^|\n)\s*内容标签\s*[:：]?\s*", t[start:])
    end = start + m_stop.start() if m_stop else len(t)
    raw = t[start:end].strip()
    out["文案"] = raw if raw else None

    return out




def parse_onebook_fields(html: str) -> Dict[str, Optional[str]]:
    soup = BeautifulSoup(html, "lxml")
    full_text = soup.get_text("\n", strip=True)

    # ✅ 临时调试：打印包含“文案”附近的片段
    t = normalize(full_text)

    # ✅ 只在 DEBUG_SNIPPET=True 时才打印
    if DEBUG_SNIPPET:
        idx = t.find("文案")
        if idx != -1:
            print("\n--- DEBUG 文案附近片段 ---")
            print(t[max(0, idx-200): idx+800])
            print("--- END DEBUG ---\n")
        else:
            print("\n[DEBUG] 页面里没出现“文案”二字\n")

    return extract_fields_from_text(full_text)



def read_rows_and_links_from_excel(xlsx_path: str) -> List[Dict]:
    """
    读取 Excel 各 sheet（跳过“晋江标签总结”），提取每行已有字段 + 作品单元格超链接
    默认列结构：A作者 B作品 C类型 D进度 E字数 F积分 G发表时间
    """
    wb = openpyxl.load_workbook(xlsx_path)

    rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() == "晋江标签总结":
            continue

        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            author = ws.cell(r, 1).value
            title = ws.cell(r, 2).value
            genre = ws.cell(r, 3).value
            status = ws.cell(r, 4).value
            word_count = ws.cell(r, 5).value
            score = ws.cell(r, 6).value
            pub_time = ws.cell(r, 7).value

            link = None
            c = ws.cell(r, 2)  # 作品列
            if c.hyperlink and c.hyperlink.target:
                link = c.hyperlink.target.strip()

            if link and "onebook.php" in link and "novelid=" in link:
                rows.append({
                    "sheet": sheet_name,
                    "作者": author,
                    "作品": title,
                    "类型": genre,
                    "进度": status,
                    "字数": word_count,
                    "积分": score,
                    "发表时间": pub_time,
                    "详情页链接": link
                })
    return rows


def main():
    base_rows = read_rows_and_links_from_excel(INPUT_XLSX)
    print(f"Found {len(base_rows)} rows with onebook links.")

    # # ✅ 只测试前 10 本
    #base_rows = base_rows[:10]
    # print(f"Testing first {len(base_rows)} books.")


    # ✅ 去重：按链接去重（保持顺序）
    seen = set()
    uniq_rows = []
    for r in base_rows:
        link = r.get("详情页链接")
        if not link:
            continue
        if link in seen:
            continue
        seen.add(link)
        uniq_rows.append(r)
    base_rows = uniq_rows
    print(f"After dedup: {len(base_rows)} rows.")

    import os
    if os.path.exists(OUTPUT_XLSX):
        old_df = pd.read_excel(OUTPUT_XLSX)
        # 兼容旧文件没有这一列的情况
        if "详情页链接" in old_df.columns:
            done_links = set(old_df["详情页链接"].dropna().astype(str))
        else:
            done_links = set()
        print(f"Resume mode: {len(done_links)} already done.")
    else:
        old_df = None
        done_links = set()

    session = requests.Session()
    results = []

    # ✅ 每处理多少本就保存一次（防止中途断掉丢进度）
    SAVE_EVERY = 200
    processed = 0
    skipped = 0

    for row in tqdm(base_rows, desc="Crawling all books"):
        link = row["详情页链接"]

        if link in done_links:
            skipped += 1
            continue

        sleep_a_bit()

        try:
            html = fetch_html(session, link)
            extra = parse_onebook_fields(html)
            row.update({
                "文案": extra.get("文案"),
                "内容标签": extra.get("内容标签"),
                "内容标签_list": extra.get("内容标签_list"),
                "一句话简介": extra.get("一句话简介"),
                "立意": extra.get("立意"),
                "error": None
            })

        except Exception as e:
            row.update({
                "文案": None,
                "内容标签": None,
                "内容标签_list": None,
                "一句话简介": None,
                "立意": None,
                "error": repr(e)
            })

        results.append(row)
        processed += 1

        # ✅ 定期落盘
        if processed % SAVE_EVERY == 0:
            df_new = pd.DataFrame(results)
            if old_df is not None and len(old_df) > 0:
                df = pd.concat([old_df, df_new], ignore_index=True)
            else:
                df = df_new

            # 清洗所有字符串列里的非法字符，避免 openpyxl 报错
            df = df.applymap(strip_illegal_xlsx_chars)

            df.to_excel(OUTPUT_XLSX, index=False)
            df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

            # 更新 done_links，避免本次运行内重复
            done_links.update(df_new["详情页链接"].dropna().astype(str))

            # 清空 results，减少内存占用
            old_df = df
            results = []

    # ✅ 循环结束后，把剩余未保存的也写进去
    if results:
        df_new = pd.DataFrame(results)
        if old_df is not None and len(old_df) > 0:
            df = pd.concat([old_df, df_new], ignore_index=True)
        else:
            df = df_new
        # 清洗所有字符串列里的非法字符，避免 openpyxl 报错
        df = df.applymap(strip_illegal_xlsx_chars)
        df.to_excel(OUTPUT_XLSX, index=False)
        df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")

    print(f"Done. processed={processed}, skipped={skipped}")
    print(f"Saved: {OUTPUT_XLSX}")
    print(f"Saved: {OUTPUT_CSV}")
    print("文案缺失本数：", df["文案"].isna().sum())
    print("内容标签缺失本数：", df["内容标签"].isna().sum())


if __name__ == "__main__":
    main()

