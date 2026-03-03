#jjwxc_crawl_csvfirst.py
#========= 步骤二，此code用来从步骤一生成的总报告里里爬取“文案”“内容标签”“一句话简介”“立意”等字段，并生成csv ========#
import re
import time
import random
from typing import Optional, Dict, List
import os

import requests
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from tqdm import tqdm

# ========= 配置 =========
# 输入的总 XLSX 报告文件（由 jj_web_multi_year.py 生成）
INPUT_XLSX = "jj_2024.xlsx"
#INPUT_XLSX = "jjwxc_2016_2025_all.xlsx"

# ✅ 运行中只写 CSV（append），最后再生成 xlsx（为了运行速度我只让这个文档生成csv）
OUTPUT_CSV = "jjwxc_2024_withtags.csv"
#OUTPUT_CSV = "jjwxc_10yrs_withtags.csv"
#FINAL_XLSX = "jjwxc_10yrs_withtags.xlsx"

#最佳休眠空隙必须 >= 0.2， 避免被封
SLEEP_MIN = 0.2
SLEEP_MAX = 0.8
RETRIES = 2
TIMEOUT = 12

DEBUG_SNIPPET = False

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://www.jjwxc.net/",
}

# Excel 不允许的控制字符（0x00-0x08,0x0B-0x0C,0x0E-0x1F）
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def strip_illegal_xlsx_chars(x):
    if isinstance(x, str):
        return _ILLEGAL_XLSX_RE.sub("", x)
    return x

# ========= ✅ 内容标签清洗 =========
VALID_TAGS = {
    "甜文","情有独钟","爽文","穿越时空","天作之合","强强","穿书","天之骄子","系统","成长","豪门世家","都市","日常","宫廷侯爵",
    "种田文","重生","仙侠修真","年代文","业界精英","破镜重圆","先婚后爱","升级流","娱乐圈","万人迷","快穿","灵异神怪","无限流",
    "幻想空间","校园","励志","基建","救赎","群像","沙雕","ABO","美食","欢喜冤家","年下","治愈","现代架空","末世","综漫",
    "追爱火葬场","星际","HE","团宠","青梅竹马","逆袭","异能","直播","暗恋","因缘邂逅","生子","萌宠","悬疑推理","美强惨",
    "高岭之花","囤货","相爱相杀","正剧","市井生活","少年漫","朝堂","西方罗曼","轻松","咒回","经营","打脸","历史衍生","柯南",
    "英美衍生","女强","荒野求生","惊悚","狗血","未来架空","女配","体育竞技","克苏鲁","抽奖抽卡","文野","三教九流","钓系","玄学",
    "近水楼台","迪化流","婚恋","单元文","马甲文","游戏网游","超级英雄","开挂","脑洞","花季雨季","布衣生活","科幻","萌娃","东方玄幻",
    "西幻","清穿","白月光","废土","爆笑","复仇虐渣","异世大陆","边缘恋歌","反套路","恋爱合约","古代幻想","日久生情","虫族","江湖",
    "论坛体","机甲","家教","鬼灭","女扮男装","魔幻","火影","科举","排球少年","乙女向","第四天灾","天选之子","阴差阳错","随身空间",
    "足球","网王","电竞","平步青云","日韩泰","龙傲天","忠犬","港风","前世今生","综艺","宅斗","赛博朋克","武侠","创业","热血","田园",
    "制服情缘","全息","规则怪谈","古穿今","失忆","腹黑","海贼王","真假少爷","御姐","权谋","宫斗","虐文","炮灰","宋穿","学霸","灵魂转换",
    "异想天开","读心术","都市异闻","咸鱼","师徒","乔装改扮","吐槽","异闻传说","时代奇缘","古典名著","剧透","姐弟恋","唐穿","史诗奇幻",
    "汉穿","哨向","男配","红楼梦","猎人","对照组","职场","时代新风","卡牌","赶山赶海","刀剑乱舞","多重人格","真假千金","现实","明穿","燃",
    "弹幕","西方名著","签到流","吐槽役","星穹铁道","秦穿","灵气复苏","总裁","位面","神话传说","转生","预知","女尊","原神","黑篮","神豪流",
    "三国穿越","高智商","犬夜叉","公路文","非遗","NPC","纸片人","少女漫","民国","大冒险","时尚圈","剑网3","群穿","毒舌","冰山","国风幻想",
    "模拟器","读档流","性别转换","FGO","傲娇","替身","烧脑","召唤流","商战","美娱","极品亲戚","吃货","封神","洪荒","开荒","奇谭","七五",
    "app","漫穿","JOJO","银魂","齐神","蓝锁","网红","暖男","萌","中二","聊斋","骑士与剑","血族","中世纪","亡灵异族","原始社会","恶役","御兽",
    "七年之痒","天降","盲盒","魔法少女","蒸汽朋克","锦鲤","扶贫","亚人","特摄","交换人生","魔王勇者","BE","死神","悲剧","红包群","网配","曲艺",
    "对话体","港台","SD","婆媳","圣斗士","绝区零"
}

_TAG_SEP_RE = re.compile(r"[、，,;；/|｜#\u3000\s]+")
_TAG_STRIP_RE = re.compile(r"^[【\[\(（<《『「]+|[】\]\)）>》』」]+$")

def clean_tags(raw: Optional[str], valid_tags: set) -> Dict[str, Optional[object]]:
    if not raw:
        return {
            "内容标签_raw": raw,
            "内容标签_clean": None,
            "内容标签_list": None,
            "内容标签_unknown": None,
        }

    s = raw.strip()
    noise_phrases = [
        "标签", "内容标签", "更多搜索", "点击", "收藏", "评论", "作者", "作品简评",
        "文章基本信息", "主角", "配角", "立意", "一句话简介"
    ]
    for w in noise_phrases:
        s = s.replace(w, " ")

    s = s.replace("：", " ").replace(":", " ").replace("#", " ")
    parts = [p for p in _TAG_SEP_RE.split(s) if p]

    cleaned, unknown, seen = [], [], set()
    for p in parts:
        p = _TAG_STRIP_RE.sub("", p.strip()).strip()
        if not p:
            continue
        if p in valid_tags:
            if p not in seen:
                cleaned.append(p)
                seen.add(p)
        else:
            if p not in seen:
                unknown.append(p)
                seen.add(p)

    return {
        "内容标签_raw": raw,
        "内容标签_clean": " ".join(cleaned) if cleaned else None,
        "内容标签_list": cleaned if cleaned else None,
        "内容标签_unknown": unknown if unknown else None,
    }

def sleep_a_bit():
    time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

def fetch_html(session: requests.Session, url: str) -> str:
    last_err = None
    for _ in range(RETRIES):
        try:
            resp = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            if not resp.encoding or resp.encoding.lower() in ("iso-8859-1", "ascii"):
                resp.encoding = resp.apparent_encoding or "gb18030"

            if resp.status_code == 200 and resp.text:
                return resp.text

            last_err = f"HTTP {resp.status_code}"
        except Exception as e:
            last_err = repr(e)

        sleep_a_bit()

    raise RuntimeError(f"Failed to fetch {url}: {last_err}")

def normalize(s: str) -> str:
    s = re.sub(r"\r", "", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def extract_fields_from_text(full_text: str) -> Dict[str, Optional[str]]:
    t = normalize(full_text)

    out = {
        "文案": None,
        "内容标签": None,
        "内容标签_list": None,
        "一句话简介": None,
        "立意": None,
    }

    lines = [x.strip() for x in t.split("\n")]

    STOP_LABELS = {
        "文案", "内容标签", "一句话简介", "立意",
        "主角", "配角", "其他", "其它", "角色",
        "作品简评", "文章基本信息", "作者", "评论",
        "所属系列", "文章进度", "全文字数", "是否出版",
        "主角视角", "配角视角"
    }

    def is_label_line(s: str) -> bool:
        s2 = s.strip()
        if not s2:
            return False
        head = s2.split("：", 1)[0].split(":", 1)[0].strip()
        return head in STOP_LABELS

    def pick_block(label: str) -> Optional[str]:
        for i, line in enumerate(lines):
            if line == label or line.startswith(label + "：") or line.startswith(label + ":"):
                vals = []
                if "：" in line:
                    tail = line.split("：", 1)[1].strip()
                    if tail:
                        vals.append(tail)
                elif ":" in line:
                    tail = line.split(":", 1)[1].strip()
                    if tail:
                        vals.append(tail)

                j = i + 1
                while j < len(lines):
                    nxt = lines[j].strip()
                    if not nxt:
                        j += 1
                        continue
                    if is_label_line(nxt) and not (nxt == label):
                        break
                    vals.append(nxt)
                    j += 1

                joined = " ".join([v for v in vals if v]).strip()
                return joined if joined else None
        return None

    out["内容标签"] = pick_block("内容标签")
    out["一句话简介"] = pick_block("一句话简介")
    out["立意"] = pick_block("立意")

    tag_pack = clean_tags(out["内容标签"], VALID_TAGS)
    out["内容标签"] = tag_pack["内容标签_clean"]
    out["内容标签_list"] = tag_pack["内容标签_list"]
    out["内容标签_raw"] = tag_pack["内容标签_raw"]
    out["内容标签_unknown"] = tag_pack["内容标签_unknown"]

    start_pat = r"(?:^|\n)\s*[【\[]?\s*文案\s*[】\]]?\s*[:：]?\s*\n?"
    m = re.search(start_pat, t)
    if not m:
        return out

    start = m.end()
    m_stop = re.search(r"(?:^|\n)\s*内容标签\s*[:：]?\s*", t[start:])
    end = start + m_stop.start() if m_stop else len(t)
    raw = t[start:end].strip()
    out["文案"] = raw if raw else None

    return out

def parse_onebook_fields(html: str) -> Dict[str, Optional[str]]:
    soup = BeautifulSoup(html, "lxml")
    full_text = soup.get_text("\n", strip=True)

    if DEBUG_SNIPPET:
        t = normalize(full_text)
        idx = t.find("文案")
        if idx != -1:
            print("\n--- DEBUG 文案附近片段 ---")
            print(t[max(0, idx-200): idx+800])
            print("--- END DEBUG ---\n")
        else:
            print("\n[DEBUG] 页面里没出现“文案”二字\n")

    return extract_fields_from_text(full_text)

def read_rows_and_links_from_excel(xlsx_path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(xlsx_path)

    rows = []
    for sheet_name in wb.sheetnames:
        if sheet_name.strip() == "晋江标签总结":
            continue

        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            author = ws.cell(r, 1).value
            title = ws.cell(r, 2).value
            genre = ws.cell(r, 3).value
            status = ws.cell(r, 4).value
            word_count = ws.cell(r, 5).value
            score = ws.cell(r, 6).value
            pub_time = ws.cell(r, 7).value

            link = None
            c = ws.cell(r, 2)
            if c.hyperlink and c.hyperlink.target:
                link = c.hyperlink.target.strip()

            if link and "onebook.php" in link and "novelid=" in link:
                rows.append({
                    "sheet": sheet_name,
                    "作者": author,
                    "作品": title,
                    "类型": genre,
                    "进度": status,
                    "字数": word_count,
                    "积分": score,
                    "发表时间": pub_time,
                    "详情页链接": link
                })
    return rows

def append_to_csv(df_new: pd.DataFrame, csv_path: str):
    """追加写入 CSV：首次写入带 header，后续 append 不写 header"""
    write_header = not os.path.exists(csv_path)
    df_new.to_csv(
        csv_path,
        mode="a",
        header=write_header,
        index=False,
        encoding="utf-8-sig"
    )

def load_done_links_from_csv(csv_path: str) -> set:
    """从现有 CSV 读取已完成链接用于断点续爬"""
    if not os.path.exists(csv_path):
        return set()
    try:
        df = pd.read_csv(csv_path, usecols=["详情页链接"], encoding="utf-8-sig")
        return set(df["详情页链接"].dropna().astype(str))
    except Exception:
        #df = pd.read_csv(csv_path, encoding="utf-8-sig")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", on_bad_lines="skip")

        if "详情页链接" in df.columns:
            return set(df["详情页链接"].dropna().astype(str))
        return set()

def export_csv_to_xlsx(csv_path: str, xlsx_path: str):
    """最后把 CSV 转成 XLSX，并清洗 Excel 非法字符（不再用 applymap，避免 FutureWarning）"""
    df = pd.read_csv(csv_path, encoding="utf-8-sig")

    # 仅对字符串列清洗非法字符
    obj_cols = df.select_dtypes(include="object").columns
    for col in obj_cols:
        df[col] = df[col].map(strip_illegal_xlsx_chars)

    # list/None 列如果有，转成字符串更好看（可选）
    for col in ["内容标签_list", "内容标签_unknown"]:
        if col in df.columns:
            df[col] = df[col].astype(str)

    df.to_excel(xlsx_path, index=False)

def main():
    base_rows = read_rows_and_links_from_excel(INPUT_XLSX)
    print(f"Found {len(base_rows)} rows with onebook links.")

    # ✅ 去重：按链接去重（保持顺序）
    seen = set()
    uniq_rows = []
    for r in base_rows:
        link = r.get("详情页链接")
        if not link or link in seen:
            continue
        seen.add(link)
        uniq_rows.append(r)
    base_rows = uniq_rows
    print(f"After dedup: {len(base_rows)} rows.")

    # ✅ Resume：从 CSV 读取 done_links（如果 CSV 已存在）
    done_links = load_done_links_from_csv(OUTPUT_CSV)
    print(f"Resume mode (CSV): {len(done_links)} already done.")

    session = requests.Session()
    results = []

    SAVE_EVERY = 200  # CSV append 很快，你可设 200~1000
    processed = 0
    skipped = 0

    try:
        pbar = tqdm(
            base_rows,
            desc="Crawling all books",
            leave=False,
            mininterval=1.0,
            dynamic_ncols=True,  # ✅ 自动适配终端宽度，避免断行
            ascii=True,          # ✅ 强烈推荐：用纯 ASCII 进度条，mac 终端最稳
            smoothing=0.1,       # ✅ 更平滑
        )
        for row in pbar:
        
            link = row["详情页链接"]

            if link in done_links:
                skipped += 1
                continue

            sleep_a_bit()

            try:
                html = fetch_html(session, link)
                extra = parse_onebook_fields(html)
                row.update({
                    "文案": extra.get("文案"),
                    "内容标签": extra.get("内容标签"),
                    "内容标签_list": extra.get("内容标签_list"),
                    "一句话简介": extra.get("一句话简介"),
                    "立意": extra.get("立意"),
                    "内容标签_raw": extra.get("内容标签_raw"),
                    "内容标签_unknown": extra.get("内容标签_unknown"),
                    "error": None
                })
            except Exception as e:
                row.update({
                    "文案": None,
                    "内容标签": None,
                    "内容标签_list": None,
                    "一句话简介": None,
                    "立意": None,
                    "内容标签_raw": None,
                    "内容标签_unknown": None,
                    "error": repr(e)
                })

            results.append(row)
            processed += 1

            # ✅ 定期追加写入 CSV（不写 xlsx）
            if processed % SAVE_EVERY == 0:
                df_new = pd.DataFrame(results)
                append_to_csv(df_new, OUTPUT_CSV)
                done_links.update(df_new["详情页链接"].dropna().astype(str))
                results = []

    except KeyboardInterrupt:
        print("\n[INFO] 检测到 Ctrl+C，中断爬取，正在保存当前内存数据到 CSV...")

    finally:
        # ✅ 把剩余未保存的也追加写进去
        if results:
            df_new = pd.DataFrame(results)
            append_to_csv(df_new, OUTPUT_CSV)
            done_links.update(df_new["详情页链接"].dropna().astype(str))
            results = []

        print(f"[INFO] Done. processed={processed}, skipped={skipped}")
        print(f"[INFO] Saved CSV: {OUTPUT_CSV}")

        # ✅ 可选：每次运行结束都生成/更新 xlsx（如果你觉得慢，可注释掉这两行）
        # print("[INFO] Exporting CSV to XLSX (final)...")
        # export_csv_to_xlsx(OUTPUT_CSV, FINAL_XLSX)
        # print(f"[INFO] Saved XLSX: {FINAL_XLSX}")

if __name__ == "__main__":
    main()
