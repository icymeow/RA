#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
jjwxc_crawl_links_only.py

只从 Excel 里的“详情页链接”爬取字段：
- 内容标签 (clean 后)
- 内容标签_list (Python list -> 写入 Excel 时转成字符串)
- 一句话简介
- 立意
- error

并写回同一个 Excel（openpyxl 原地更新），支持断点续爬。
"""

import re
import time
import random
from typing import Optional, Dict

import requests
import openpyxl
from bs4 import BeautifulSoup
from tqdm import tqdm

# ========= 配置 =========
INPUT_XLSX = "jj_2024.xlsx"        # 你的新表
OUTPUT_XLSX = "jj_2024_filled.xlsx"  # 建议写到新文件，避免覆盖损坏；想覆盖就改成同名

SLEEP_MIN = 0.2
SLEEP_MAX = 0.8
RETRIES = 2
TIMEOUT = 12
SAVE_EVERY = 50  # 每处理 N 行保存一次，防止意外中断丢数据

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Referer": "https://www.jjwxc.net/",
}

# Excel 不允许的控制字符（0x00-0x08,0x0B-0x0C,0x0E-0x1F）
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

def strip_illegal_xlsx_chars(x):
    if isinstance(x, str):
        return _ILLEGAL_XLSX_RE.sub("", x)
    return x

# ========= 内容标签清洗 =========
VALID_TAGS = {
    "甜文","情有独钟","爽文","穿越时空","天作之合","强强","穿书","天之骄子","系统","成长","豪门世家","都市","日常","宫廷侯爵",
    "种田文","重生","仙侠修真","年代文","业界精英","破镜重圆","先婚后爱","升级流","娱乐圈","万人迷","快穿","灵异神怪","无限流",
    "幻想空间","校园","励志","基建","救赎","群像","沙雕","ABO","美食","欢喜冤家","年下","治愈","现代架空","末世","综漫",
    "追爱火葬场","星际","HE","团宠","青梅竹马","逆袭","异能","直播","暗恋","因缘邂逅","生子","萌宠","悬疑推理","美强惨",
    "高岭之花","囤货","相爱相杀","正剧","市井生活","少年漫","朝堂","西方罗曼","轻松","咒回","经营","打脸","历史衍生","柯南",
    "英美衍生","女强","荒野求生","惊悚","狗血","未来架空","女配","体育竞技","克苏鲁","抽奖抽卡","文野","三教九流","钓系","玄学",
    "近水楼台","迪化流","婚恋","单元文","马甲文","游戏网游","超级英雄","开挂","脑洞","花季雨季","布衣生活","科幻","萌娃","东方玄幻",
    "西幻","清穿","白月光","废土","爆笑","复仇虐渣","异世大陆","边缘恋歌","反套路","恋爱合约","古代幻想","日久生情","虫族","江湖",
    "论坛体","机甲","家教","鬼灭","女扮男装","魔幻","火影","科举","排球少年","乙女向","第四天灾","天选之子","阴差阳错","随身空间",
    "足球","网王","电竞","平步青云","日韩泰","龙傲天","忠犬","港风","前世今生","综艺","宅斗","赛博朋克","武侠","创业","热血","田园",
    "制服情缘","全息","规则怪谈","古穿今","失忆","腹黑","海贼王","真假少爷","御姐","权谋","宫斗","虐文","炮灰","宋穿","学霸","灵魂转换",
    "异想天开","读心术","都市异闻","咸鱼","师徒","乔装改扮","吐槽","异闻传说","时代奇缘","古典名著","剧透","姐弟恋","唐穿","史诗奇幻",
    "汉穿","哨向","男配","红楼梦","猎人","对照组","职场","时代新风","卡牌","赶山赶海","刀剑乱舞","多重人格","真假千金","现实","明穿","燃",
    "弹幕","西方名著","签到流","吐槽役","星穹铁道","秦穿","灵气复苏","总裁","位面","神话传说","转生","预知","女尊","原神","黑篮","神豪流",
    "三国穿越","高智商","犬夜叉","公路文","非遗","NPC","纸片人","少女漫","民国","大冒险","时尚圈","剑网3","群穿","毒舌","冰山","国风幻想",
    "模拟器","读档流","性别转换","FGO","傲娇","替身","烧脑","召唤流","商战","美娱","极品亲戚","吃货","封神","洪荒","开荒","奇谭","七五",
    "app","漫穿","JOJO","银魂","齐神","蓝锁","网红","暖男","萌","中二","聊斋","骑士与剑","血族","中世纪","亡灵异族","原始社会","恶役","御兽",
    "七年之痒","天降","盲盒","魔法少女","蒸汽朋克","锦鲤","扶贫","亚人","特摄","交换人生","魔王勇者","BE","死神","悲剧","红包群","网配","曲艺",
    "对话体","港台","SD","婆媳","圣斗士","绝区零"
}

_TAG_SEP_RE = re.compile(r"[、，,;；/|｜#\u3000\s]+")
_TAG_STRIP_RE = re.compile(r"^[【\[\(（<《『「]+|[】\]\)）>》』」]+$")

def clean_tags(raw: Optional[str], valid_tags: set) -> Dict[str, Optional[object]]:
    if not raw:
        return {"clean": None, "lst": None, "raw": raw, "unknown": None}

    s = raw.strip()
    noise_phrases = [
        "标签", "内容标签", "更多搜索", "点击", "收藏", "评论", "作者", "作品简评",
        "文章基本信息", "主角", "配角", "立意", "一句话简介"
    ]
    for w in noise_phrases:
        s = s.replace(w, " ")

    s = s.replace("：", " ").replace(":", " ").replace("#", " ")
    parts = [p for p in _TAG_SEP_RE.split(s) if p]

    cleaned, unknown, seen = [], [], set()
    for p in parts:
        p = _TAG_STRIP_RE.sub("", p.strip()).strip()
        if not p:
            continue
        if p in valid_tags:
            if p not in seen:
                cleaned.append(p)
                seen.add(p)
        else:
            if p not in seen:
                unknown.append(p)
                seen.add(p)

    return {
        "clean": " ".join(cleaned) if cleaned else None,
        "lst": cleaned if cleaned else None,
        "raw": raw,
        "unknown": unknown if unknown else None,
    }

def sleep_a_bit():
    time.sleep(random.uniform(SLEEP_MIN, SLEEP_MAX))

def fetch_html(session: requests.Session, url: str) -> str:
    last_err = None
    for _ in range(RETRIES):
        try:
            resp = session.get(url, headers=HEADERS, timeout=TIMEOUT)
            if not resp.encoding or resp.encoding.lower() in ("iso-8859-1", "ascii"):
                resp.encoding = resp.apparent_encoding or "gb18030"
            if resp.status_code == 200 and resp.text:
                return resp.text
            last_err = f"HTTP {resp.status_code}"
        except Exception as e:
            last_err = repr(e)
        sleep_a_bit()
    raise RuntimeError(f"Failed to fetch {url}: {last_err}")

def normalize(s: str) -> str:
    s = re.sub(r"\r", "", s)
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def extract_needed_fields(full_text: str) -> Dict[str, Optional[object]]:
    """
    从页面纯文本中提取：
    - 内容标签 (clean)
    - 内容标签_list (list)
    - 一句话简介
    - 立意
    """
    t = normalize(full_text)
    lines = [x.strip() for x in t.split("\n")]

    STOP_LABELS = {
        "内容标签", "一句话简介", "立意",
        "主角", "配角", "其他", "其它", "角色",
        "作品简评", "文章基本信息", "作者", "评论",
        "所属系列", "文章进度", "全文字数", "是否出版",
        "主角视角", "配角视角", "文案"
    }

    def is_label_line(s: str) -> bool:
        s2 = s.strip()
        if not s2:
            return False
        head = s2.split("：", 1)[0].split(":", 1)[0].strip()
        return head in STOP_LABELS

    def pick_block(label: str) -> Optional[str]:
        for i, line in enumerate(lines):
            if line == label or line.startswith(label + "：") or line.startswith(label + ":"):
                vals = []
                if "：" in line:
                    tail = line.split("：", 1)[1].strip()
                    if tail:
                        vals.append(tail)
                elif ":" in line:
                    tail = line.split(":", 1)[1].strip()
                    if tail:
                        vals.append(tail)

                j = i + 1
                while j < len(lines):
                    nxt = lines[j].strip()
                    if not nxt:
                        j += 1
                        continue
                    if is_label_line(nxt) and nxt != label:
                        break
                    vals.append(nxt)
                    j += 1

                joined = " ".join([v for v in vals if v]).strip()
                return joined if joined else None
        return None

    raw_tags = pick_block("内容标签")
    one_line = pick_block("一句话简介")
    theme = pick_block("立意")

    tag_pack = clean_tags(raw_tags, VALID_TAGS)

    return {
        "内容标签": tag_pack["clean"],
        "内容标签_list": tag_pack["lst"],
        "一句话简介": one_line,
        "立意": theme,
        "内容标签_raw": tag_pack["raw"],          # 你不需要可删除
        "内容标签_unknown": tag_pack["unknown"],  # 你不需要可删除
    }

def parse_onebook(html: str) -> Dict[str, Optional[object]]:
    soup = BeautifulSoup(html, "lxml")
    full_text = soup.get_text("\n", strip=True)
    return extract_needed_fields(full_text)

def find_col_indices(ws) -> Dict[str, int]:
    """
    从第一行表头找列号（1-based），要求至少有：详情页链接 / 内容标签 / 内容标签_list / 一句话简介 / 立意 / error
    """
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and v.strip():
            header[v.strip()] = c

    required = ["详情页链接", "内容标签", "内容标签_list", "一句话简介", "立意", "error"]
    missing = [x for x in required if x not in header]
    if missing:
        raise ValueError(f"Missing columns in header row: {missing}")
    return header

def is_row_done(ws, r: int, col: Dict[str, int]) -> bool:
    """
    断点续爬策略：只要这行里（内容标签/一句话简介/立意/error）任意一个已有值，就认为“处理过”，跳过。
    你也可以改成更严格：必须四个都非空才算 done。
    """
    for k in ["内容标签", "一句话简介", "立意", "error"]:
        v = ws.cell(r, col[k]).value
        if v is not None and str(v).strip() != "":
            return True
    return False

def main():
    wb = openpyxl.load_workbook(INPUT_XLSX)
    # 默认处理第一个 sheet；如果你要指定 sheet 名字，改这里：
    ws = wb[wb.sheetnames[0]]

    col = find_col_indices(ws)

    # 统计需要处理的行
    targets = []
    for r in range(2, ws.max_row + 1):
        link = ws.cell(r, col["详情页链接"]).value
        if not link or not isinstance(link, str):
            continue
        link = link.strip()
        if "onebook.php" not in link or "novelid=" not in link:
            continue
        if is_row_done(ws, r, col):
            continue
        targets.append((r, link))

    print(f"Total rows: {ws.max_row - 1}, to crawl: {len(targets)}")

    session = requests.Session()
    processed = 0

    try:
        for (r, link) in tqdm(targets, desc="Crawling links", dynamic_ncols=True, ascii=True):
            sleep_a_bit()
            try:
                html = fetch_html(session, link)
                extra = parse_onebook(html)

                tags_clean = strip_illegal_xlsx_chars(extra.get("内容标签"))
                tags_list = extra.get("内容标签_list")
                one_line = strip_illegal_xlsx_chars(extra.get("一句话简介"))
                theme = strip_illegal_xlsx_chars(extra.get("立意"))

                # 写回 Excel
                ws.cell(r, col["内容标签"]).value = tags_clean
                ws.cell(r, col["内容标签_list"]).value = str(tags_list) if tags_list else None
                ws.cell(r, col["一句话简介"]).value = one_line
                ws.cell(r, col["立意"]).value = theme
                ws.cell(r, col["error"]).value = None

            except Exception as e:
                ws.cell(r, col["内容标签"]).value = None
                ws.cell(r, col["内容标签_list"]).value = None
                ws.cell(r, col["一句话简介"]).value = None
                ws.cell(r, col["立意"]).value = None
                ws.cell(r, col["error"]).value = strip_illegal_xlsx_chars(repr(e))

            processed += 1
            if processed % SAVE_EVERY == 0:
                wb.save(OUTPUT_XLSX)

    except KeyboardInterrupt:
        print("\n[INFO] Ctrl+C detected, saving...")

    finally:
        wb.save(OUTPUT_XLSX)
        print(f"[INFO] Saved: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()