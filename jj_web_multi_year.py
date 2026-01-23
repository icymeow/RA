#步骤一：用这个文件爬取 jjwxc 网站上指定年份范围内的基础数据，保存为 JSONL 和 XLSX 文件。

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# jj_web_multi_year.py
# 这个文件是用来爬取 JJWXC 网站上指定年份范围内的小说数据的脚本。
# 它会根据用户提供的年份范围，逐年爬取字段：
# 作者、作品（带超链接）、类型、进度、字数、作品积分、发表时间
# 并将每年的数据保存为单独的 JSONL 文件和 XLSX 文件，可在out文件夹里看到每年的结果。
# 最后，还可以将所有年份的数据合并为一个总的 XLSX 文件，方便用户进行整体分析和查看。
# 生成的文档名称格式如：jjwxc_2016_top50pages.xlsx、...、jjwxc_2016_2025_all.xlsx

#如果想运行此脚本，可以参考以下命令行,示例的是2016年到2025年榜单前50页小说（一页100本）的爬取命令：
'''
python3 jj_web_multi_year.py \
  --start-year 2016 \
  --end-year 2025 \
  --pages-per-year 50 \
  --cookie "$(cat cookie.txt)"
'''
  

import argparse
import json
import random
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Set
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

#晋江网站
BASE_URL = "https://www.jjwxc.net/bookbase.php"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.7",
    "Cache-Control": "no-cache",
    "Pragma": "no-cache",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Referer": BASE_URL,
}


def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()


def robust_get(session: requests.Session, params: dict, retries: int = 6, timeout: int = 25) -> str:
    last_err = None
    for i in range(retries):
        try:
            r = session.get(BASE_URL, params=params, headers=HEADERS, timeout=timeout)
            r.raise_for_status()
            if not r.encoding or r.encoding.lower() == "iso-8859-1":
                r.encoding = r.apparent_encoding or "gbk"
            return r.text
        except Exception as e:
            last_err = e
            time.sleep(2 + i * 2 + random.random())
    raise RuntimeError(f"请求失败，重试 {retries} 次仍失败：{last_err}")


def extract_total_pages_m_p(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, "lxml")
    inp = soup.find("input", attrs={"name": "m_p"})
    if inp and inp.get("value"):
        return inp["value"]
    text = soup.get_text(" ", strip=True)
    m = re.search(r"共\s*(\d+)\s*页", text)
    return m.group(1) if m else None


def find_target_table(soup: BeautifulSoup):
    for table in soup.find_all("table"):
        header_text = clean_text(table.get_text(" "))
        if all(k in header_text for k in ["作者", "作品", "类型", "进度", "字数", "作品积分", "发表时间"]):
            return table
    return None


def parse_rows(html: str) -> List[Dict]:
    # 明确登录拦截提示（你之前贴的）
    if "请" in html and "登入" in html and "后再访问此页面" in html:
        raise PermissionError("页面提示需登入（cookie失效/未带cookie）")

    soup = BeautifulSoup(html, "lxml")
    table = find_target_table(soup)
    if not table:
        raise ValueError("未找到目标表格（可能反爬/需要登录/结构变化）")

    rows: List[Dict] = []
    for tr in table.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) < 7:
            continue

        author = clean_text(tds[0].get_text(" "))

        a = tds[1].find("a")
        work_title = clean_text(a.get_text(" ")) if a else clean_text(tds[1].get_text(" "))
        work_href = urljoin(BASE_URL, a["href"]) if a and a.get("href") else ""

        category = clean_text(tds[2].get_text(" "))
        progress = clean_text(tds[3].get_text(" "))
        word_count = clean_text(tds[4].get_text(" "))
        score = clean_text(tds[5].get_text(" "))
        pub_time = clean_text(tds[6].get_text(" "))

        if author == "作者" and work_title == "作品":
            continue
        if not author or not work_title:
            continue

        rows.append({
            "作者": author,
            "作品": work_title,
            "作品链接": work_href,
            "类型": category,
            "进度": progress,
            "字数": word_count,
            "作品积分": score,
            "发表时间": pub_time,
        })
    return rows


def base_params_for_year(year: int) -> dict:
    """
    JJWXC 这里年份参数形如：fbsj2022=2022
    """
    return {
        "fw0": "0",
        f"fbsj{year}": str(year),
        "novelbefavoritedcount0": "0",
        "yc0": "0",
        "xx0": "0",
        "mainview0": "0",
        "sd0": "0",
        "lx0": "0",
        "collectiontypes": "ors",
        "notlikecollectiontypes": "ors",
        "bq": "-1",
        "removebq": "",
        "searchkeywords": "",
        "sortType": "0",
        "isfinish": "0",
    }


def build_params(year_params: dict, page: int, m_p: Optional[str]) -> dict:
    p = dict(year_params)
    p["page"] = str(page)
    if m_p:
        p["m_p"] = str(m_p)
    return p


def load_done_pages(jsonl_path: Path) -> Set[int]:
    done: Set[int] = set()
    if not jsonl_path.exists():
        return done
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
                pg = obj.get("_page")
                if isinstance(pg, int):
                    done.add(pg)
            except Exception:
                continue
    return done


def append_page_rows(jsonl_path: Path, page: int, year: int, rows: List[Dict]):
    with jsonl_path.open("a", encoding="utf-8") as f:
        for r in rows:
            r2 = dict(r)
            r2["_page"] = page
            r2["_year"] = year
            f.write(json.dumps(r2, ensure_ascii=False) + "\n")


def load_all_rows(jsonl_path: Path) -> List[Dict]:
    rows: List[Dict] = []
    if not jsonl_path.exists():
        return rows
    with jsonl_path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            obj.pop("_page", None)
            obj.pop("_year", None)
            rows.append(obj)
    return rows


def save_to_xlsx(rows: List[Dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "JJWXC"

    headers = ["作者", "作品", "类型", "进度", "字数", "作品积分", "发表时间"]
    ws.append(headers)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r.get("作者", ""),
            r.get("作品", ""),
            r.get("类型", ""),
            r.get("进度", ""),
            r.get("字数", ""),
            r.get("作品积分", ""),
            r.get("发表时间", ""),
        ])
        link = r.get("作品链接", "")
        if link:
            c = ws.cell(row=ws.max_row, column=2)
            c.hyperlink = link
            c.font = Font(color="0000EE", underline="single")

    for col in range(1, len(headers) + 1):
        max_len = 10
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(row=rr, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    wb.save(out_path)


def crawl_one_year(
    session: requests.Session,
    year: int,
    pages_per_year: int,
    jsonl_path: Path,
    year_out_xlsx: Path,
    debug_dir: Path,
    failed_log: Path,
    min_delay: float,
    max_delay: float,
    fail_sleep_min: float,
    fail_sleep_max: float,
):
    year_params = base_params_for_year(year)
    debug_dir.mkdir(parents=True, exist_ok=True)
    failed_log.parent.mkdir(parents=True, exist_ok=True)

    done_pages = load_done_pages(jsonl_path)
    need_pages = [p for p in range(1, pages_per_year + 1) if p not in done_pages]
    print(f"\n===== Year {year} =====")
    print(f"📌 已完成页数：{len(done_pages)} / {pages_per_year}，待爬：{len(need_pages)} 页")

    # 先拿第一页用于 m_p
    first_html = robust_get(session, build_params(year_params, 1, None))
    m_p = extract_total_pages_m_p(first_html)

    for p in tqdm(need_pages, desc=f"Year {year}", leave=True):
        html = ""
        try:
            html = first_html if p == 1 else robust_get(session, build_params(year_params, p, m_p))
            rows = parse_rows(html)

            if len(rows) < 50:
                raise RuntimeError(f"解析条数异常偏少：{len(rows)}")

            append_page_rows(jsonl_path, p, year, rows)
            if p % 10 == 0:
                time.sleep(random.uniform(15, 35))
            #time.sleep(random.uniform(min_delay, max_delay))

        except Exception as e:
            # 保存失败 html
            html_path = debug_dir / f"page_{p}_failed.html"
            try:
                html_path.write_text(html or "", encoding="utf-8", errors="ignore")
            except Exception:
                pass

            with failed_log.open("a", encoding="utf-8") as f:
                f.write(f"{p}\t{type(e).__name__}: {e}\n")

            print(f"\n⚠️ {year} 第 {p} 页失败：{type(e).__name__}: {e}")
            print(f"   - 已保存：{html_path}")
            print(f"   - 已记录：{failed_log}\n")

            # 如果明确是登录问题，别继续刷了（保护账号/避免加重风控）
            if isinstance(e, PermissionError):
                print("❗检测到需登入页面：请更新 cookie 后再继续。已停止本年份爬取。")
                break

            time.sleep(random.uniform(fail_sleep_min, fail_sleep_max))
            continue

    # 年度导出
    rows = load_all_rows(jsonl_path)
    save_to_xlsx(rows, year_out_xlsx)
    print(f"✅ Year {year} 导出：{year_out_xlsx}（累计 {len(rows)} 条）")


def merge_years_to_one_xlsx(years: List[int], cache_dir: Path, out_path: Path):
    all_rows: List[Dict] = []
    for y in years:
        jsonl = cache_dir / f"jjwxc_{y}.jsonl"
        if jsonl.exists():
            rows = load_all_rows(jsonl)
            # 给总表加一个“年份”列（可选：不想要可删）
            for r in rows:
                r["年份"] = str(y)
            all_rows.extend(rows)

    # 输出总表（把“年份”放最前）
    wb = Workbook()
    ws = wb.active
    ws.title = "JJWXC_ALL"

    headers = ["年份", "作者", "作品", "类型", "进度", "字数", "作品积分", "发表时间"]
    ws.append(headers)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in all_rows:
        ws.append([
            r.get("年份", ""),
            r.get("作者", ""),
            r.get("作品", ""),
            r.get("类型", ""),
            r.get("进度", ""),
            r.get("字数", ""),
            r.get("作品积分", ""),
            r.get("发表时间", ""),
        ])
        link = r.get("作品链接", "")
        if link:
            c = ws.cell(row=ws.max_row, column=3)  # “作品”在第3列
            c.hyperlink = link
            c.font = Font(color="0000EE", underline="single")

    for col in range(1, len(headers) + 1):
        max_len = 10
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(row=rr, column=col).value
            if v:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--start-year", type=int, default=2014)
    ap.add_argument("--end-year", type=int, default=2022)
    ap.add_argument("--pages-per-year", type=int, default=50)

    ap.add_argument("--cookie", type=str, default=None, help="建议用：--cookie \"$(cat cookie.txt)\"")
    ap.add_argument("--min-delay", type=float, default=4.0)
    ap.add_argument("--max-delay", type=float, default=10.0)
    ap.add_argument("--fail-sleep-min", type=float, default=10.0)
    ap.add_argument("--fail-sleep-max", type=float, default=20.0)

    ap.add_argument("--cache-dir", type=str, default="cache")
    ap.add_argument("--out-dir", type=str, default="out")
    ap.add_argument("--debug-dir", type=str, default="debug_html")
    ap.add_argument("--failed-dir", type=str, default="failed_pages")
    ap.add_argument("--no-merge", action="store_true", help="不生成总合并表")

    args = ap.parse_args()

    cache_dir = Path(args.cache_dir); cache_dir.mkdir(exist_ok=True)
    out_dir = Path(args.out_dir); out_dir.mkdir(exist_ok=True)

    debug_root = Path(args.debug_dir)
    failed_root = Path(args.failed_dir)

    years = list(range(args.end_year, args.start_year - 1, -1))  # 2022 -> 2014

    with requests.Session() as session:
        if args.cookie:
            session.headers.update({"Cookie": args.cookie})

        for y in years:
            jsonl_path = cache_dir / f"jjwxc_{y}.jsonl"
            year_xlsx = out_dir / f"jjwxc_{y}_top{args.pages_per_year}pages.xlsx"
            debug_dir = debug_root / str(y)
            failed_log = failed_root / f"{y}_failed_pages.txt"

            crawl_one_year(
                session=session,
                year=y,
                pages_per_year=args.pages_per_year,
                jsonl_path=jsonl_path,
                year_out_xlsx=year_xlsx,
                debug_dir=debug_dir,
                failed_log=failed_log,
                min_delay=args.min_delay,
                max_delay=args.max_delay,
                fail_sleep_min=args.fail_sleep_min,
                fail_sleep_max=args.fail_sleep_max,
            )

    if not args.no_merge:
        merged = out_dir / f"jjwxc_{args.start_year}_{args.end_year}_all.xlsx"
        merge_years_to_one_xlsx(years, cache_dir, merged)
        print(f"\n✅ 合并总表导出：{merged}")


if __name__ == "__main__":
    main()
